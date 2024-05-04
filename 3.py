from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime, timedelta
import os

# Load workbook 11.xlsx and get Sheet3
wb_11 = load_workbook(filename='11.xlsx', data_only=True)
ws_11 = wb_11['Sheet3']

# Find the first Monday in the date column of Sheet3
date_column_11 = ws_11['A']
first_monday_row_11 = None
for cell in date_column_11:
    if isinstance(cell.value, datetime) and cell.value.weekday() == 0:  # Check if the cell contains a datetime object
        first_monday_row_11 = cell.row
        break

if first_monday_row_11 is None:
    raise ValueError("No Monday found in the date column of Sheet3 in 11.xlsx")

# Extract the dates from row 2 and row 29 of column A
date_row_2 = date_column_11[first_monday_row_11].value.date()  # Extract only the date part
date_row_29 = date_column_11[first_monday_row_11 + 27].value.date()  # Extract only the date part

# Create a new workbook 33.xlsx
wb_33 = Workbook()
ws_33 = wb_33.active
ws_33.title = 'Sheet1'

# Load workbook 22.xlsx and get Sheet1
wb_22 = load_workbook(filename='22.xlsx', read_only=True)
ws_22 = wb_22['Sheet1']

# Copy the whole Sheet1 of 22.xlsx to 33.xlsx with formatting
for row in ws_22.iter_rows(min_row=1, max_row=ws_22.max_row, min_col=1, max_col=ws_22.max_column):
    for cell in row:
        if cell.value is not None:
            copied_cell = ws_33.cell(row=cell.row, column=cell.column, value=cell.value)

            # Copy cell formatting
            if cell.has_style:
                copied_cell.font = Font(name=cell.font.name,
                                        size=cell.font.size,
                                        bold=cell.font.bold,
                                        italic=cell.font.italic,
                                        vertAlign=cell.font.vertAlign,
                                        underline=cell.font.underline,
                                        strike=cell.font.strike,
                                        color=cell.font.color)
                copied_cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                                  vertical=cell.alignment.vertical,
                                                  wrap_text=cell.alignment.wrap_text,
                                                  shrink_to_fit=cell.alignment.shrink_to_fit,
                                                  indent=cell.alignment.indent)
                copied_cell.number_format = cell.number_format
                copied_cell.protection = cell.protection
                copied_cell.border = cell.border
                copied_cell.fill = cell.fill

# Fill the first column of the new workbook from row 1 to row 28 with values from the second Monday
for i in range(first_monday_row_11 + 56, first_monday_row_11 + 84):
    date_value = ws_11.cell(row=i, column=1).value.date()  # Extract only the date part
    ws_33.cell(row=i - (first_monday_row_11 + 56) + 2, column=1, value=date_value)

    # Fill column E of the new workbook with values from column B of Sheet3 of 11.xlsx
    value_B = ws_11.cell(row=i, column=2).value
    ws_33.cell(row=i - (first_monday_row_11 + 56) + 2, column=5, value=value_B)

# Extract the dates from 33.xlsx
date_column_33 = ws_33['A']
date_row_2 = date_column_33[1].value  # Extract the date
date_row_29 = date_column_33[28].value  # Extract the date

# Save the workbook as 33.xlsx
filename = f"{date_row_2.strftime('%Y-%m-%d')}~{date_row_29.strftime('%Y-%m-%d')}.xlsx"
wb_33.save(filename)

# Check if the original 33.xlsx exists before renaming it
if os.path.exists('33.xlsx'):
    os.remove('33.xlsx')
    print("Original file 33.xlsx has been deleted.")

print(f"File {filename} has been created.")
