import warnings
import openpyxl

# Filter out the warning about header/footer parsing
warnings.filterwarnings("ignore", category=UserWarning)

# Load the workbook
wb = openpyxl.load_workbook('11.xlsx')

# Select the original sheet
original_sheet = wb['Sheet1']

# Create a new sheet
new_sheet = wb.create_sheet(title='Sheet2')

# Iterate through rows and extract values from the original sheet
for row_index, row in enumerate(original_sheet.iter_rows(values_only=True), start=1):
    # Extract the first part (date) from the first column
    date_part = row[0].split()[0]
    # Write the date part to the first column of the new sheet
    new_sheet.cell(row=row_index, column=1, value=date_part)

    # Extract the value from the second column
    second_column_value = row[1]
    # Check if the cell in the second column is not empty
    if second_column_value:
        # Split the value into two parts
        second_column_parts = second_column_value.split()
        first_part = second_column_parts[0]
        second_part = second_column_parts[1] if len(second_column_parts) > 1 else None
    else:
        first_part = None
        second_part = None

    # Write the first part to the second column of the new sheet
    new_sheet.cell(row=row_index, column=2, value=first_part)

    # Write the second part to the third column of the new sheet
    new_sheet.cell(row=row_index, column=3, value=second_part)

# Save the workbook (overwrite the original file)
wb.save('11.xlsx')

import pandas as pd

# Load data from the source Excel file
df1 = pd.read_excel('11.xlsx', sheet_name='Sheet2')

# Generate a date range from April 1st to June 30th, repeating each date 4 times
dates = pd.date_range(start='2024-04-01', end='2024-06-30', freq='D').repeat(4)

# Create a new DataFrame for Sheet2
df2 = pd.DataFrame({'日期': dates})

# Map values from Sheet1 to Sheet2
strings = []
for i, date in enumerate(dates):
    row_idx = (i // 4) % len(df1)
    col_idx = 1 if (i % 4 == 0 or i % 4 == 2) else 2
    strings.append(df1.iloc[row_idx, col_idx])

df2['字符串'] = strings

# Convert datetime column to date only
df2['日期'] = df2['日期'].dt.date

# Save the DataFrame to the second sheet of the Excel file
with pd.ExcelWriter('11.xlsx', mode='a', engine='openpyxl') as writer:
    df2.to_excel(writer, sheet_name='Sheet3', index=False)

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime, timedelta
import os

# Load workbook 11.xlsx and get Sheet3
wb_11 = load_workbook(filename='11.xlsx', data_only=True)
ws_11 = wb_11['Sheet3']

# Find the first Monday in the date column of Sheet3
date_column_11 = ws_11['A']
first_monday_row_11 = None
for cell in date_column_11:
    if isinstance(cell.value, datetime) and cell.value.weekday() == 0:  # Check if the cell contains a datetime object
        first_monday_row_11 = cell.row
        break

if first_monday_row_11 is None:
    raise ValueError("No Monday found in the date column of Sheet3 in 11.xlsx")

# Extract the dates from row 2 and row 29 of column A
date_row_2 = date_column_11[first_monday_row_11].value.date()  # Extract only the date part
date_row_29 = date_column_11[first_monday_row_11 + 27].value.date()  # Extract only the date part

# Create a new workbook 33.xlsx
wb_33 = Workbook()
ws_33 = wb_33.active
ws_33.title = 'Sheet1'

# Load workbook 22.xlsx and get Sheet1
wb_22 = load_workbook(filename='22.xlsx', read_only=True)
ws_22 = wb_22['Sheet1']

# Copy the whole Sheet1 of 22.xlsx to 33.xlsx with formatting
for row in ws_22.iter_rows(min_row=1, max_row=ws_22.max_row, min_col=1, max_col=ws_22.max_column):
    for cell in row:
        if cell.value is not None:
            copied_cell = ws_33.cell(row=cell.row, column=cell.column, value=cell.value)

            # Copy cell formatting
            if cell.has_style:
                copied_cell.font = Font(name=cell.font.name,
                                        size=cell.font.size,
                                        bold=cell.font.bold,
                                        italic=cell.font.italic,
                                        vertAlign=cell.font.vertAlign,
                                        underline=cell.font.underline,
                                        strike=cell.font.strike,
                                        color=cell.font.color)
                copied_cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                                  vertical=cell.alignment.vertical,
                                                  wrap_text=cell.alignment.wrap_text,
                                                  shrink_to_fit=cell.alignment.shrink_to_fit,
                                                  indent=cell.alignment.indent)
                copied_cell.number_format = cell.number_format
                copied_cell.protection = cell.protection
                copied_cell.border = cell.border
                copied_cell.fill = cell.fill

# Fill the first column of 33.xlsx from row 2 to row 29 with values from the first Monday to 27 rows below the first Monday
for i in range(first_monday_row_11, first_monday_row_11 + 28):
    date_value = ws_11.cell(row=i, column=1).value.date()  # Extract only the date part
    ws_33.cell(row=i - first_monday_row_11 + 2, column=1, value=date_value)

    # Fill column E of 33.xlsx with values from column B of Sheet3 of 11.xlsx
    value_B = ws_11.cell(row=i, column=2).value
    ws_33.cell(row=i - first_monday_row_11 + 2, column=5, value=value_B)

# Save the workbook as 33.xlsx
filename = f"{date_row_2.strftime('%Y-%m-%d')}~{date_row_29.strftime('%Y-%m-%d')}.xlsx"
wb_33.save(filename)

# Check if the original 33.xlsx exists before renaming it
if os.path.exists('33.xlsx'):
    os.remove('33.xlsx')
    print("Original file 33.xlsx has been deleted.")

print(f"File {filename} has been created.")

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime, timedelta
import os

# Load workbook 11.xlsx and get Sheet3
wb_11 = load_workbook(filename='11.xlsx', data_only=True)
ws_11 = wb_11['Sheet3']

# Find the first Monday in the date column of Sheet3
date_column_11 = ws_11['A']
first_monday_row_11 = None
for cell in date_column_11:
    if isinstance(cell.value, datetime) and cell.value.weekday() == 0:  # Check if the cell contains a datetime object
        first_monday_row_11 = cell.row
        break

if first_monday_row_11 is None:
    raise ValueError("No Monday found in the date column of Sheet3 in 11.xlsx")

# Extract the dates from row 2 and row 29 of column A
date_row_2 = date_column_11[first_monday_row_11].value.date()  # Extract only the date part
date_row_29 = date_column_11[first_monday_row_11 + 27].value.date()  # Extract only the date part

# Create a new workbook 33.xlsx
wb_33 = Workbook()
ws_33 = wb_33.active
ws_33.title = 'Sheet1'

# Load workbook 22.xlsx and get Sheet1
wb_22 = load_workbook(filename='22.xlsx', read_only=True)
ws_22 = wb_22['Sheet1']

# Copy the whole Sheet1 of 22.xlsx to 33.xlsx with formatting
for row in ws_22.iter_rows(min_row=1, max_row=ws_22.max_row, min_col=1, max_col=ws_22.max_column):
    for cell in row:
        if cell.value is not None:
            copied_cell = ws_33.cell(row=cell.row, column=cell.column, value=cell.value)

            # Copy cell formatting
            if cell.has_style:
                copied_cell.font = Font(name=cell.font.name,
                                        size=cell.font.size,
                                        bold=cell.font.bold,
                                        italic=cell.font.italic,
                                        vertAlign=cell.font.vertAlign,
                                        underline=cell.font.underline,
                                        strike=cell.font.strike,
                                        color=cell.font.color)
                copied_cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                                  vertical=cell.alignment.vertical,
                                                  wrap_text=cell.alignment.wrap_text,
                                                  shrink_to_fit=cell.alignment.shrink_to_fit,
                                                  indent=cell.alignment.indent)
                copied_cell.number_format = cell.number_format
                copied_cell.protection = cell.protection
                copied_cell.border = cell.border
                copied_cell.fill = cell.fill

# Fill the first column of the new workbook from row 1 to row 28 with values from the second Monday
for i in range(first_monday_row_11 + 28, first_monday_row_11 + 56):
    date_value = ws_11.cell(row=i, column=1).value.date()  # Extract only the date part
    ws_33.cell(row=i - (first_monday_row_11 + 28) + 2, column=1, value=date_value)

    # Fill column E of the new workbook with values from column B of Sheet3 of 11.xlsx
    value_B = ws_11.cell(row=i, column=2).value
    ws_33.cell(row=i - (first_monday_row_11 + 28) + 2, column=5, value=value_B)

# Extract the dates from 33.xlsx
date_column_33 = ws_33['A']
date_row_2 = date_column_33[1].value  # Extract the date
date_row_29 = date_column_33[28].value  # Extract the date

# Save the workbook as 33.xlsx
filename = f"{date_row_2.strftime('%Y-%m-%d')}~{date_row_29.strftime('%Y-%m-%d')}.xlsx"
wb_33.save(filename)

# Check if the original 33.xlsx exists before renaming it
if os.path.exists('33.xlsx'):
    os.remove('33.xlsx')
    print("Original file 33.xlsx has been deleted.")

print(f"File {filename} has been created.")

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime, timedelta
import os

# Load workbook 11.xlsx and get Sheet3
wb_11 = load_workbook(filename='11.xlsx', data_only=True)
ws_11 = wb_11['Sheet3']

# Find the first Monday in the date column of Sheet3
date_column_11 = ws_11['A']
first_monday_row_11 = None
for cell in date_column_11:
    if isinstance(cell.value, datetime) and cell.value.weekday() == 0:  # Check if the cell contains a datetime object
        first_monday_row_11 = cell.row
        break

if first_monday_row_11 is None:
    raise ValueError("No Monday found in the date column of Sheet3 in 11.xlsx")

# Extract the dates from row 2 and row 29 of column A
date_row_2 = date_column_11[first_monday_row_11].value.date()  # Extract only the date part
date_row_29 = date_column_11[first_monday_row_11 + 27].value.date()  # Extract only the date part

# Create a new workbook 33.xlsx
wb_33 = Workbook()
ws_33 = wb_33.active
ws_33.title = 'Sheet1'

# Load workbook 22.xlsx and get Sheet1
wb_22 = load_workbook(filename='22.xlsx', read_only=True)
ws_22 = wb_22['Sheet1']

# Copy the whole Sheet1 of 22.xlsx to 33.xlsx with formatting
for row in ws_22.iter_rows(min_row=1, max_row=ws_22.max_row, min_col=1, max_col=ws_22.max_column):
    for cell in row:
        if cell.value is not None:
            copied_cell = ws_33.cell(row=cell.row, column=cell.column, value=cell.value)

            # Copy cell formatting
            if cell.has_style:
                copied_cell.font = Font(name=cell.font.name,
                                        size=cell.font.size,
                                        bold=cell.font.bold,
                                        italic=cell.font.italic,
                                        vertAlign=cell.font.vertAlign,
                                        underline=cell.font.underline,
                                        strike=cell.font.strike,
                                        color=cell.font.color)
                copied_cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                                  vertical=cell.alignment.vertical,
                                                  wrap_text=cell.alignment.wrap_text,
                                                  shrink_to_fit=cell.alignment.shrink_to_fit,
                                                  indent=cell.alignment.indent)
                copied_cell.number_format = cell.number_format
                copied_cell.protection = cell.protection
                copied_cell.border = cell.border
                copied_cell.fill = cell.fill

# Fill the first column of the new workbook from row 1 to row 28 with values from the second Monday
for i in range(first_monday_row_11 + 56, first_monday_row_11 + 84):
    date_value = ws_11.cell(row=i, column=1).value.date()  # Extract only the date part
    ws_33.cell(row=i - (first_monday_row_11 + 56) + 2, column=1, value=date_value)

    # Fill column E of the new workbook with values from column B of Sheet3 of 11.xlsx
    value_B = ws_11.cell(row=i, column=2).value
    ws_33.cell(row=i - (first_monday_row_11 + 56) + 2, column=5, value=value_B)

# Extract the dates from 33.xlsx
date_column_33 = ws_33['A']
date_row_2 = date_column_33[1].value  # Extract the date
date_row_29 = date_column_33[28].value  # Extract the date

# Save the workbook as 33.xlsx
filename = f"{date_row_2.strftime('%Y-%m-%d')}~{date_row_29.strftime('%Y-%m-%d')}.xlsx"
wb_33.save(filename)

# Check if the original 33.xlsx exists before renaming it
if os.path.exists('33.xlsx'):
    os.remove('33.xlsx')
    print("Original file 33.xlsx has been deleted.")

print(f"File {filename} has been created.")

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime, timedelta
import os

# Load workbook 11.xlsx and get Sheet3
wb_11 = load_workbook(filename='11.xlsx', data_only=True)
ws_11 = wb_11['Sheet3']

# Find the first Monday in the date column of Sheet3
date_column_11 = ws_11['A']
first_monday_row_11 = None
for cell in date_column_11:
    if isinstance(cell.value, datetime) and cell.value.weekday() == 0:  # Check if the cell contains a datetime object
        first_monday_row_11 = cell.row
        break

if first_monday_row_11 is None:
    raise ValueError("No Monday found in the date column of Sheet3 in 11.xlsx")

# Extract the dates from row 2 and row 29 of column A
date_row_2 = date_column_11[first_monday_row_11].value.date()  # Extract only the date part
date_row_29 = date_column_11[first_monday_row_11 + 27].value.date()  # Extract only the date part

# Create a new workbook 33.xlsx
wb_33 = Workbook()
ws_33 = wb_33.active
ws_33.title = 'Sheet1'

# Load workbook 22.xlsx and get Sheet1
wb_22 = load_workbook(filename='22.xlsx', read_only=True)
ws_22 = wb_22['Sheet1']

# Copy the whole Sheet1 of 22.xlsx to 33.xlsx with formatting
for row in ws_22.iter_rows(min_row=1, max_row=ws_22.max_row, min_col=1, max_col=ws_22.max_column):
    for cell in row:
        if cell.value is not None:
            copied_cell = ws_33.cell(row=cell.row, column=cell.column, value=cell.value)

            # Copy cell formatting
            if cell.has_style:
                copied_cell.font = Font(name=cell.font.name,
                                        size=cell.font.size,
                                        bold=cell.font.bold,
                                        italic=cell.font.italic,
                                        vertAlign=cell.font.vertAlign,
                                        underline=cell.font.underline,
                                        strike=cell.font.strike,
                                        color=cell.font.color)
                copied_cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                                  vertical=cell.alignment.vertical,
                                                  wrap_text=cell.alignment.wrap_text,
                                                  shrink_to_fit=cell.alignment.shrink_to_fit,
                                                  indent=cell.alignment.indent)
                copied_cell.number_format = cell.number_format
                copied_cell.protection = cell.protection
                copied_cell.border = cell.border
                copied_cell.fill = cell.fill

# Fill the first column of the new workbook from row 1 to row 28 with values from the second Monday
for i in range(first_monday_row_11 + 84, first_monday_row_11 + 112):
    date_value = ws_11.cell(row=i, column=1).value.date()  # Extract only the date part
    ws_33.cell(row=i - (first_monday_row_11 + 84) + 2, column=1, value=date_value)

    # Fill column E of the new workbook with values from column B of Sheet3 of 11.xlsx
    value_B = ws_11.cell(row=i, column=2).value
    ws_33.cell(row=i - (first_monday_row_11 + 84) + 2, column=5, value=value_B)

# Extract the dates from 33.xlsx
date_column_33 = ws_33['A']
date_row_2 = date_column_33[1].value  # Extract the date
date_row_29 = date_column_33[28].value  # Extract the date

# Save the workbook as 33.xlsx
filename = f"{date_row_2.strftime('%Y-%m-%d')}~{date_row_29.strftime('%Y-%m-%d')}.xlsx"
wb_33.save(filename)

# Check if the original 33.xlsx exists before renaming it
if os.path.exists('33.xlsx'):
    os.remove('33.xlsx')
    print("Original file 33.xlsx has been deleted.")

print(f"File {filename} has been created.")

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime, timedelta
import os

# Load workbook 11.xlsx and get Sheet3
wb_11 = load_workbook(filename='11.xlsx', data_only=True)
ws_11 = wb_11['Sheet3']

# Find the first Monday in the date column of Sheet3
date_column_11 = ws_11['A']
first_monday_row_11 = None
for cell in date_column_11:
    if isinstance(cell.value, datetime) and cell.value.weekday() == 0:  # Check if the cell contains a datetime object
        first_monday_row_11 = cell.row
        break

if first_monday_row_11 is None:
    raise ValueError("No Monday found in the date column of Sheet3 in 11.xlsx")

# Extract the dates from row 2 and row 29 of column A
date_row_2 = date_column_11[first_monday_row_11].value.date()  # Extract only the date part
date_row_29 = date_column_11[first_monday_row_11 + 27].value.date()  # Extract only the date part

# Create a new workbook 33.xlsx
wb_33 = Workbook()
ws_33 = wb_33.active
ws_33.title = 'Sheet1'

# Load workbook 22.xlsx and get Sheet1
wb_22 = load_workbook(filename='22.xlsx', read_only=True)
ws_22 = wb_22['Sheet1']

# Copy the whole Sheet1 of 22.xlsx to 33.xlsx with formatting
for row in ws_22.iter_rows(min_row=1, max_row=ws_22.max_row, min_col=1, max_col=ws_22.max_column):
    for cell in row:
        if cell.value is not None:
            copied_cell = ws_33.cell(row=cell.row, column=cell.column, value=cell.value)

            # Copy cell formatting
            if cell.has_style:
                copied_cell.font = Font(name=cell.font.name,
                                        size=cell.font.size,
                                        bold=cell.font.bold,
                                        italic=cell.font.italic,
                                        vertAlign=cell.font.vertAlign,
                                        underline=cell.font.underline,
                                        strike=cell.font.strike,
                                        color=cell.font.color)
                copied_cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                                  vertical=cell.alignment.vertical,
                                                  wrap_text=cell.alignment.wrap_text,
                                                  shrink_to_fit=cell.alignment.shrink_to_fit,
                                                  indent=cell.alignment.indent)
                copied_cell.number_format = cell.number_format
                copied_cell.protection = cell.protection
                copied_cell.border = cell.border
                copied_cell.fill = cell.fill

# Fill the first column of the new workbook from row 1 to row 28 with values from the second Monday
for i in range(first_monday_row_11 + 112, first_monday_row_11 + 140):
    date_value = ws_11.cell(row=i, column=1).value.date()  # Extract only the date part
    ws_33.cell(row=i - (first_monday_row_11 + 112) + 2, column=1, value=date_value)

    # Fill column E of the new workbook with values from column B of Sheet3 of 11.xlsx
    value_B = ws_11.cell(row=i, column=2).value
    ws_33.cell(row=i - (first_monday_row_11 + 112) + 2, column=5, value=value_B)

# Extract the dates from 33.xlsx
date_column_33 = ws_33['A']
date_row_2 = date_column_33[1].value  # Extract the date
date_row_29 = date_column_33[28].value  # Extract the date

# Save the workbook as 33.xlsx
filename = f"{date_row_2.strftime('%Y-%m-%d')}~{date_row_29.strftime('%Y-%m-%d')}.xlsx"
wb_33.save(filename)

# Check if the original 33.xlsx exists before renaming it
if os.path.exists('33.xlsx'):
    os.remove('33.xlsx')
    print("Original file 33.xlsx has been deleted.")

print(f"File {filename} has been created.")

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime, timedelta
import os

# Load workbook 11.xlsx and get Sheet3
wb_11 = load_workbook(filename='11.xlsx', data_only=True)
ws_11 = wb_11['Sheet3']

# Find the first Monday in the date column of Sheet3
date_column_11 = ws_11['A']
first_monday_row_11 = None
for cell in date_column_11:
    if isinstance(cell.value, datetime) and cell.value.weekday() == 0:  # Check if the cell contains a datetime object
        first_monday_row_11 = cell.row
        break

if first_monday_row_11 is None:
    raise ValueError("No Monday found in the date column of Sheet3 in 11.xlsx")

# Extract the dates from row 2 and row 29 of column A
date_row_2 = date_column_11[first_monday_row_11].value.date()  # Extract only the date part
date_row_29 = date_column_11[first_monday_row_11 + 27].value.date()  # Extract only the date part

# Create a new workbook 33.xlsx
wb_33 = Workbook()
ws_33 = wb_33.active
ws_33.title = 'Sheet1'

# Load workbook 22.xlsx and get Sheet1
wb_22 = load_workbook(filename='22.xlsx', read_only=True)
ws_22 = wb_22['Sheet1']

# Copy the whole Sheet1 of 22.xlsx to 33.xlsx with formatting
for row in ws_22.iter_rows(min_row=1, max_row=ws_22.max_row, min_col=1, max_col=ws_22.max_column):
    for cell in row:
        if cell.value is not None:
            copied_cell = ws_33.cell(row=cell.row, column=cell.column, value=cell.value)

            # Copy cell formatting
            if cell.has_style:
                copied_cell.font = Font(name=cell.font.name,
                                        size=cell.font.size,
                                        bold=cell.font.bold,
                                        italic=cell.font.italic,
                                        vertAlign=cell.font.vertAlign,
                                        underline=cell.font.underline,
                                        strike=cell.font.strike,
                                        color=cell.font.color)
                copied_cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                                  vertical=cell.alignment.vertical,
                                                  wrap_text=cell.alignment.wrap_text,
                                                  shrink_to_fit=cell.alignment.shrink_to_fit,
                                                  indent=cell.alignment.indent)
                copied_cell.number_format = cell.number_format
                copied_cell.protection = cell.protection
                copied_cell.border = cell.border
                copied_cell.fill = cell.fill

# Fill the first column of the new workbook from row 1 to row 28 with values from the second Monday
for i in range(first_monday_row_11 + 140, first_monday_row_11 + 168):
    date_value = ws_11.cell(row=i, column=1).value.date()  # Extract only the date part
    ws_33.cell(row=i - (first_monday_row_11 + 140) + 2, column=1, value=date_value)

    # Fill column E of the new workbook with values from column B of Sheet3 of 11.xlsx
    value_B = ws_11.cell(row=i, column=2).value
    ws_33.cell(row=i - (first_monday_row_11 + 140) + 2, column=5, value=value_B)

# Extract the dates from 33.xlsx
date_column_33 = ws_33['A']
date_row_2 = date_column_33[1].value  # Extract the date
date_row_29 = date_column_33[28].value  # Extract the date

# Save the workbook as 33.xlsx
filename = f"{date_row_2.strftime('%Y-%m-%d')}~{date_row_29.strftime('%Y-%m-%d')}.xlsx"
wb_33.save(filename)

# Check if the original 33.xlsx exists before renaming it
if os.path.exists('33.xlsx'):
    os.remove('33.xlsx')
    print("Original file 33.xlsx has been deleted.")

print(f"File {filename} has been created.")

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime, timedelta
import os

# Load workbook 11.xlsx and get Sheet3
wb_11 = load_workbook(filename='11.xlsx', data_only=True)
ws_11 = wb_11['Sheet3']

# Find the first Monday in the date column of Sheet3
date_column_11 = ws_11['A']
first_monday_row_11 = None
for cell in date_column_11:
    if isinstance(cell.value, datetime) and cell.value.weekday() == 0:  # Check if the cell contains a datetime object
        first_monday_row_11 = cell.row
        break

if first_monday_row_11 is None:
    raise ValueError("No Monday found in the date column of Sheet3 in 11.xlsx")

# Extract the dates from row 2 and row 29 of column A
date_row_2 = date_column_11[first_monday_row_11].value.date()  # Extract only the date part
date_row_29 = date_column_11[first_monday_row_11 + 27].value.date()  # Extract only the date part

# Create a new workbook 33.xlsx
wb_33 = Workbook()
ws_33 = wb_33.active
ws_33.title = 'Sheet1'

# Load workbook 22.xlsx and get Sheet1
wb_22 = load_workbook(filename='22.xlsx', read_only=True)
ws_22 = wb_22['Sheet1']

# Copy the whole Sheet1 of 22.xlsx to 33.xlsx with formatting
for row in ws_22.iter_rows(min_row=1, max_row=ws_22.max_row, min_col=1, max_col=ws_22.max_column):
    for cell in row:
        if cell.value is not None:
            copied_cell = ws_33.cell(row=cell.row, column=cell.column, value=cell.value)

            # Copy cell formatting
            if cell.has_style:
                copied_cell.font = Font(name=cell.font.name,
                                        size=cell.font.size,
                                        bold=cell.font.bold,
                                        italic=cell.font.italic,
                                        vertAlign=cell.font.vertAlign,
                                        underline=cell.font.underline,
                                        strike=cell.font.strike,
                                        color=cell.font.color)
                copied_cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                                  vertical=cell.alignment.vertical,
                                                  wrap_text=cell.alignment.wrap_text,
                                                  shrink_to_fit=cell.alignment.shrink_to_fit,
                                                  indent=cell.alignment.indent)
                copied_cell.number_format = cell.number_format
                copied_cell.protection = cell.protection
                copied_cell.border = cell.border
                copied_cell.fill = cell.fill

# Fill the first column of the new workbook from row 1 to row 28 with values from the second Monday
for i in range(first_monday_row_11 + 168, first_monday_row_11 + 196):
    date_value = ws_11.cell(row=i, column=1).value.date()  # Extract only the date part
    ws_33.cell(row=i - (first_monday_row_11 + 168) + 2, column=1, value=date_value)

    # Fill column E of the new workbook with values from column B of Sheet3 of 11.xlsx
    value_B = ws_11.cell(row=i, column=2).value
    ws_33.cell(row=i - (first_monday_row_11 + 168) + 2, column=5, value=value_B)

# Extract the dates from 33.xlsx
date_column_33 = ws_33['A']
date_row_2 = date_column_33[1].value  # Extract the date
date_row_29 = date_column_33[28].value  # Extract the date

# Save the workbook as 33.xlsx
filename = f"{date_row_2.strftime('%Y-%m-%d')}~{date_row_29.strftime('%Y-%m-%d')}.xlsx"
wb_33.save(filename)

# Check if the original 33.xlsx exists before renaming it
if os.path.exists('33.xlsx'):
    os.remove('33.xlsx')
    print("Original file 33.xlsx has been deleted.")

print(f"File {filename} has been created.")

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime, timedelta
import os

# Load workbook 11.xlsx and get Sheet3
wb_11 = load_workbook(filename='11.xlsx', data_only=True)
ws_11 = wb_11['Sheet3']

# Find the first Monday in the date column of Sheet3
date_column_11 = ws_11['A']
first_monday_row_11 = None
for cell in date_column_11:
    if isinstance(cell.value, datetime) and cell.value.weekday() == 0:  # Check if the cell contains a datetime object
        first_monday_row_11 = cell.row
        break

if first_monday_row_11 is None:
    raise ValueError("No Monday found in the date column of Sheet3 in 11.xlsx")

# Extract the dates from row 2 and row 29 of column A
date_row_2 = date_column_11[first_monday_row_11].value.date()  # Extract only the date part
date_row_29 = date_column_11[first_monday_row_11 + 27].value.date()  # Extract only the date part

# Create a new workbook 33.xlsx
wb_33 = Workbook()
ws_33 = wb_33.active
ws_33.title = 'Sheet1'

# Load workbook 22.xlsx and get Sheet1
wb_22 = load_workbook(filename='22.xlsx', read_only=True)
ws_22 = wb_22['Sheet1']

# Copy the whole Sheet1 of 22.xlsx to 33.xlsx with formatting
for row in ws_22.iter_rows(min_row=1, max_row=ws_22.max_row, min_col=1, max_col=ws_22.max_column):
    for cell in row:
        if cell.value is not None:
            copied_cell = ws_33.cell(row=cell.row, column=cell.column, value=cell.value)

            # Copy cell formatting
            if cell.has_style:
                copied_cell.font = Font(name=cell.font.name,
                                        size=cell.font.size,
                                        bold=cell.font.bold,
                                        italic=cell.font.italic,
                                        vertAlign=cell.font.vertAlign,
                                        underline=cell.font.underline,
                                        strike=cell.font.strike,
                                        color=cell.font.color)
                copied_cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                                  vertical=cell.alignment.vertical,
                                                  wrap_text=cell.alignment.wrap_text,
                                                  shrink_to_fit=cell.alignment.shrink_to_fit,
                                                  indent=cell.alignment.indent)
                copied_cell.number_format = cell.number_format
                copied_cell.protection = cell.protection
                copied_cell.border = cell.border
                copied_cell.fill = cell.fill

# Fill the first column of the new workbook from row 1 to row 28 with values from the second Monday
for i in range(first_monday_row_11 + 196, first_monday_row_11 + 224):
    date_value = ws_11.cell(row=i, column=1).value.date()  # Extract only the date part
    ws_33.cell(row=i - (first_monday_row_11 + 196) + 2, column=1, value=date_value)

    # Fill column E of the new workbook with values from column B of Sheet3 of 11.xlsx
    value_B = ws_11.cell(row=i, column=2).value
    ws_33.cell(row=i - (first_monday_row_11 + 196) + 2, column=5, value=value_B)

# Extract the dates from 33.xlsx
date_column_33 = ws_33['A']
date_row_2 = date_column_33[1].value  # Extract the date
date_row_29 = date_column_33[28].value  # Extract the date

# Save the workbook as 33.xlsx
filename = f"{date_row_2.strftime('%Y-%m-%d')}~{date_row_29.strftime('%Y-%m-%d')}.xlsx"
wb_33.save(filename)

# Check if the original 33.xlsx exists before renaming it
if os.path.exists('33.xlsx'):
    os.remove('33.xlsx')
    print("Original file 33.xlsx has been deleted.")

print(f"File {filename} has been created.")

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime, timedelta
import os

# Load workbook 11.xlsx and get Sheet3
wb_11 = load_workbook(filename='11.xlsx', data_only=True)
ws_11 = wb_11['Sheet3']

# Find the first Monday in the date column of Sheet3
date_column_11 = ws_11['A']
first_monday_row_11 = None
for cell in date_column_11:
    if isinstance(cell.value, datetime) and cell.value.weekday() == 0:  # Check if the cell contains a datetime object
        first_monday_row_11 = cell.row
        break

if first_monday_row_11 is None:
    raise ValueError("No Monday found in the date column of Sheet3 in 11.xlsx")

# Extract the dates from row 2 and row 29 of column A
date_row_2 = date_column_11[first_monday_row_11].value.date()  # Extract only the date part
date_row_29 = date_column_11[first_monday_row_11 + 27].value.date()  # Extract only the date part

# Create a new workbook 33.xlsx
wb_33 = Workbook()
ws_33 = wb_33.active
ws_33.title = 'Sheet1'

# Load workbook 22.xlsx and get Sheet1
wb_22 = load_workbook(filename='22.xlsx', read_only=True)
ws_22 = wb_22['Sheet1']

# Copy the whole Sheet1 of 22.xlsx to 33.xlsx with formatting
for row in ws_22.iter_rows(min_row=1, max_row=ws_22.max_row, min_col=1, max_col=ws_22.max_column):
    for cell in row:
        if cell.value is not None:
            copied_cell = ws_33.cell(row=cell.row, column=cell.column, value=cell.value)

            # Copy cell formatting
            if cell.has_style:
                copied_cell.font = Font(name=cell.font.name,
                                        size=cell.font.size,
                                        bold=cell.font.bold,
                                        italic=cell.font.italic,
                                        vertAlign=cell.font.vertAlign,
                                        underline=cell.font.underline,
                                        strike=cell.font.strike,
                                        color=cell.font.color)
                copied_cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                                  vertical=cell.alignment.vertical,
                                                  wrap_text=cell.alignment.wrap_text,
                                                  shrink_to_fit=cell.alignment.shrink_to_fit,
                                                  indent=cell.alignment.indent)
                copied_cell.number_format = cell.number_format
                copied_cell.protection = cell.protection
                copied_cell.border = cell.border
                copied_cell.fill = cell.fill

# Fill the first column of the new workbook from row 1 to row 28 with values from the second Monday
for i in range(first_monday_row_11 + 224, first_monday_row_11 + 252):
    date_value = ws_11.cell(row=i, column=1).value.date()  # Extract only the date part
    ws_33.cell(row=i - (first_monday_row_11 + 224) + 2, column=1, value=date_value)

    # Fill column E of the new workbook with values from column B of Sheet3 of 11.xlsx
    value_B = ws_11.cell(row=i, column=2).value
    ws_33.cell(row=i - (first_monday_row_11 + 224) + 2, column=5, value=value_B)

# Extract the dates from 33.xlsx
date_column_33 = ws_33['A']
date_row_2 = date_column_33[1].value  # Extract the date
date_row_29 = date_column_33[28].value  # Extract the date

# Save the workbook as 33.xlsx
filename = f"{date_row_2.strftime('%Y-%m-%d')}~{date_row_29.strftime('%Y-%m-%d')}.xlsx"
wb_33.save(filename)

# Check if the original 33.xlsx exists before renaming it
if os.path.exists('33.xlsx'):
    os.remove('33.xlsx')
    print("Original file 33.xlsx has been deleted.")

print(f"File {filename} has been created.")

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime, timedelta
import os

# Load workbook 11.xlsx and get Sheet3
wb_11 = load_workbook(filename='11.xlsx', data_only=True)
ws_11 = wb_11['Sheet3']

# Find the first Monday in the date column of Sheet3
date_column_11 = ws_11['A']
first_monday_row_11 = None
for cell in date_column_11:
    if isinstance(cell.value, datetime) and cell.value.weekday() == 0:  # Check if the cell contains a datetime object
        first_monday_row_11 = cell.row
        break

if first_monday_row_11 is None:
    raise ValueError("No Monday found in the date column of Sheet3 in 11.xlsx")

# Extract the dates from row 2 and row 29 of column A
date_row_2 = date_column_11[first_monday_row_11].value.date()  # Extract only the date part
date_row_29 = date_column_11[first_monday_row_11 + 27].value.date()  # Extract only the date part

# Create a new workbook 33.xlsx
wb_33 = Workbook()
ws_33 = wb_33.active
ws_33.title = 'Sheet1'

# Load workbook 22.xlsx and get Sheet1
wb_22 = load_workbook(filename='22.xlsx', read_only=True)
ws_22 = wb_22['Sheet1']

# Copy the whole Sheet1 of 22.xlsx to 33.xlsx with formatting
for row in ws_22.iter_rows(min_row=1, max_row=ws_22.max_row, min_col=1, max_col=ws_22.max_column):
    for cell in row:
        if cell.value is not None:
            copied_cell = ws_33.cell(row=cell.row, column=cell.column, value=cell.value)

            # Copy cell formatting
            if cell.has_style:
                copied_cell.font = Font(name=cell.font.name,
                                        size=cell.font.size,
                                        bold=cell.font.bold,
                                        italic=cell.font.italic,
                                        vertAlign=cell.font.vertAlign,
                                        underline=cell.font.underline,
                                        strike=cell.font.strike,
                                        color=cell.font.color)
                copied_cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                                  vertical=cell.alignment.vertical,
                                                  wrap_text=cell.alignment.wrap_text,
                                                  shrink_to_fit=cell.alignment.shrink_to_fit,
                                                  indent=cell.alignment.indent)
                copied_cell.number_format = cell.number_format
                copied_cell.protection = cell.protection
                copied_cell.border = cell.border
                copied_cell.fill = cell.fill

# Fill the first column of the new workbook from row 1 to row 28 with values from the second Monday
for i in range(first_monday_row_11 + 252, first_monday_row_11 + 280):
    date_value = ws_11.cell(row=i, column=1).value.date()  # Extract only the date part
    ws_33.cell(row=i - (first_monday_row_11 + 252) + 2, column=1, value=date_value)

    # Fill column E of the new workbook with values from column B of Sheet3 of 11.xlsx
    value_B = ws_11.cell(row=i, column=2).value
    ws_33.cell(row=i - (first_monday_row_11 + 252) + 2, column=5, value=value_B)

# Extract the dates from 33.xlsx
date_column_33 = ws_33['A']
date_row_2 = date_column_33[1].value  # Extract the date
date_row_29 = date_column_33[28].value  # Extract the date

# Save the workbook as 33.xlsx
filename = f"{date_row_2.strftime('%Y-%m-%d')}~{date_row_29.strftime('%Y-%m-%d')}.xlsx"
wb_33.save(filename)

# Check if the original 33.xlsx exists before renaming it
if os.path.exists('33.xlsx'):
    os.remove('33.xlsx')
    print("Original file 33.xlsx has been deleted.")

print(f"File {filename} has been created.")

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime, timedelta
import os

# Load workbook 11.xlsx and get Sheet3
wb_11 = load_workbook(filename='11.xlsx', data_only=True)
ws_11 = wb_11['Sheet3']

# Find the first Monday in the date column of Sheet3
date_column_11 = ws_11['A']
first_monday_row_11 = None
for cell in date_column_11:
    if isinstance(cell.value, datetime) and cell.value.weekday() == 0:  # Check if the cell contains a datetime object
        first_monday_row_11 = cell.row
        break

if first_monday_row_11 is None:
    raise ValueError("No Monday found in the date column of Sheet3 in 11.xlsx")

# Extract the dates from row 2 and row 29 of column A
date_row_2 = date_column_11[first_monday_row_11].value.date()  # Extract only the date part
date_row_29 = date_column_11[first_monday_row_11 + 27].value.date()  # Extract only the date part

# Create a new workbook 33.xlsx
wb_33 = Workbook()
ws_33 = wb_33.active
ws_33.title = 'Sheet1'

# Load workbook 22.xlsx and get Sheet1
wb_22 = load_workbook(filename='22.xlsx', read_only=True)
ws_22 = wb_22['Sheet1']

# Copy the whole Sheet1 of 22.xlsx to 33.xlsx with formatting
for row in ws_22.iter_rows(min_row=1, max_row=ws_22.max_row, min_col=1, max_col=ws_22.max_column):
    for cell in row:
        if cell.value is not None:
            copied_cell = ws_33.cell(row=cell.row, column=cell.column, value=cell.value)

            # Copy cell formatting
            if cell.has_style:
                copied_cell.font = Font(name=cell.font.name,
                                        size=cell.font.size,
                                        bold=cell.font.bold,
                                        italic=cell.font.italic,
                                        vertAlign=cell.font.vertAlign,
                                        underline=cell.font.underline,
                                        strike=cell.font.strike,
                                        color=cell.font.color)
                copied_cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                                  vertical=cell.alignment.vertical,
                                                  wrap_text=cell.alignment.wrap_text,
                                                  shrink_to_fit=cell.alignment.shrink_to_fit,
                                                  indent=cell.alignment.indent)
                copied_cell.number_format = cell.number_format
                copied_cell.protection = cell.protection
                copied_cell.border = cell.border
                copied_cell.fill = cell.fill

# Fill the first column of the new workbook from row 1 to row 28 with values from the second Monday
for i in range(first_monday_row_11 + 280, first_monday_row_11 + 308):
    date_value = ws_11.cell(row=i, column=1).value.date()  # Extract only the date part
    ws_33.cell(row=i - (first_monday_row_11 + 280) + 2, column=1, value=date_value)

    # Fill column E of the new workbook with values from column B of Sheet3 of 11.xlsx
    value_B = ws_11.cell(row=i, column=2).value
    ws_33.cell(row=i - (first_monday_row_11 + 280) + 2, column=5, value=value_B)

# Extract the dates from 33.xlsx
date_column_33 = ws_33['A']
date_row_2 = date_column_33[1].value  # Extract the date
date_row_29 = date_column_33[28].value  # Extract the date

# Save the workbook as 33.xlsx
filename = f"{date_row_2.strftime('%Y-%m-%d')}~{date_row_29.strftime('%Y-%m-%d')}.xlsx"
wb_33.save(filename)

# Check if the original 33.xlsx exists before renaming it
if os.path.exists('33.xlsx'):
    os.remove('33.xlsx')
    print("Original file 33.xlsx has been deleted.")

print(f"File {filename} has been created.")

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime, timedelta
import os

# Load workbook 11.xlsx and get Sheet3
wb_11 = load_workbook(filename='11.xlsx', data_only=True)
ws_11 = wb_11['Sheet3']

# Find the first Monday in the date column of Sheet3
date_column_11 = ws_11['A']
first_monday_row_11 = None
for cell in date_column_11:
    if isinstance(cell.value, datetime) and cell.value.weekday() == 0:  # Check if the cell contains a datetime object
        first_monday_row_11 = cell.row
        break

if first_monday_row_11 is None:
    raise ValueError("No Monday found in the date column of Sheet3 in 11.xlsx")

# Extract the dates from row 2 and row 29 of column A
date_row_2 = date_column_11[first_monday_row_11].value.date()  # Extract only the date part
date_row_29 = date_column_11[first_monday_row_11 + 27].value.date()  # Extract only the date part

# Create a new workbook 33.xlsx
wb_33 = Workbook()
ws_33 = wb_33.active
ws_33.title = 'Sheet1'

# Load workbook 22.xlsx and get Sheet1
wb_22 = load_workbook(filename='22.xlsx', read_only=True)
ws_22 = wb_22['Sheet1']

# Copy the whole Sheet1 of 22.xlsx to 33.xlsx with formatting
for row in ws_22.iter_rows(min_row=1, max_row=ws_22.max_row, min_col=1, max_col=ws_22.max_column):
    for cell in row:
        if cell.value is not None:
            copied_cell = ws_33.cell(row=cell.row, column=cell.column, value=cell.value)

            # Copy cell formatting
            if cell.has_style:
                copied_cell.font = Font(name=cell.font.name,
                                        size=cell.font.size,
                                        bold=cell.font.bold,
                                        italic=cell.font.italic,
                                        vertAlign=cell.font.vertAlign,
                                        underline=cell.font.underline,
                                        strike=cell.font.strike,
                                        color=cell.font.color)
                copied_cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                                  vertical=cell.alignment.vertical,
                                                  wrap_text=cell.alignment.wrap_text,
                                                  shrink_to_fit=cell.alignment.shrink_to_fit,
                                                  indent=cell.alignment.indent)
                copied_cell.number_format = cell.number_format
                copied_cell.protection = cell.protection
                copied_cell.border = cell.border
                copied_cell.fill = cell.fill

# Fill the first column of the new workbook from row 1 to row 28 with values from the second Monday
for i in range(first_monday_row_11 + 308, first_monday_row_11 + 336):
    date_value = ws_11.cell(row=i, column=1).value.date()  # Extract only the date part
    ws_33.cell(row=i - (first_monday_row_11 + 308) + 2, column=1, value=date_value)

    # Fill column E of the new workbook with values from column B of Sheet3 of 11.xlsx
    value_B = ws_11.cell(row=i, column=2).value
    ws_33.cell(row=i - (first_monday_row_11 + 308) + 2, column=5, value=value_B)

# Extract the dates from 33.xlsx
date_column_33 = ws_33['A']
date_row_2 = date_column_33[1].value  # Extract the date
date_row_29 = date_column_33[28].value  # Extract the date

# Save the workbook as 33.xlsx
filename = f"{date_row_2.strftime('%Y-%m-%d')}~{date_row_29.strftime('%Y-%m-%d')}.xlsx"
wb_33.save(filename)

# Check if the original 33.xlsx exists before renaming it
if os.path.exists('33.xlsx'):
    os.remove('33.xlsx')
    print("Original file 33.xlsx has been deleted.")

print(f"File {filename} has been created.")

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime, timedelta
import os

# Load workbook 11.xlsx and get Sheet3
wb_11 = load_workbook(filename='11.xlsx', data_only=True)
ws_11 = wb_11['Sheet3']

# Find the first Monday in the date column of Sheet3
date_column_11 = ws_11['A']
first_monday_row_11 = None
for cell in date_column_11:
    if isinstance(cell.value, datetime) and cell.value.weekday() == 0:  # Check if the cell contains a datetime object
        first_monday_row_11 = cell.row
        break

if first_monday_row_11 is None:
    raise ValueError("No Monday found in the date column of Sheet3 in 11.xlsx")

# Extract the dates from row 2 and row 29 of column A
date_row_2 = date_column_11[first_monday_row_11].value.date()  # Extract only the date part
date_row_29 = date_column_11[first_monday_row_11 + 27].value.date()  # Extract only the date part

# Create a new workbook 33.xlsx
wb_33 = Workbook()
ws_33 = wb_33.active
ws_33.title = 'Sheet1'

# Load workbook 22.xlsx and get Sheet1
wb_22 = load_workbook(filename='22.xlsx', read_only=True)
ws_22 = wb_22['Sheet1']

# Copy the whole Sheet1 of 22.xlsx to 33.xlsx with formatting
for row in ws_22.iter_rows(min_row=1, max_row=ws_22.max_row, min_col=1, max_col=ws_22.max_column):
    for cell in row:
        if cell.value is not None:
            copied_cell = ws_33.cell(row=cell.row, column=cell.column, value=cell.value)

            # Copy cell formatting
            if cell.has_style:
                copied_cell.font = Font(name=cell.font.name,
                                        size=cell.font.size,
                                        bold=cell.font.bold,
                                        italic=cell.font.italic,
                                        vertAlign=cell.font.vertAlign,
                                        underline=cell.font.underline,
                                        strike=cell.font.strike,
                                        color=cell.font.color)
                copied_cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                                  vertical=cell.alignment.vertical,
                                                  wrap_text=cell.alignment.wrap_text,
                                                  shrink_to_fit=cell.alignment.shrink_to_fit,
                                                  indent=cell.alignment.indent)
                copied_cell.number_format = cell.number_format
                copied_cell.protection = cell.protection
                copied_cell.border = cell.border
                copied_cell.fill = cell.fill

# Fill the first column of the new workbook from row 1 to row 28 with values from the second Monday
for i in range(first_monday_row_11 + 336, first_monday_row_11 + 364):
    date_value = ws_11.cell(row=i, column=1).value.date()  # Extract only the date part
    ws_33.cell(row=i - (first_monday_row_11 + 336) + 2, column=1, value=date_value)

    # Fill column E of the new workbook with values from column B of Sheet3 of 11.xlsx
    value_B = ws_11.cell(row=i, column=2).value
    ws_33.cell(row=i - (first_monday_row_11 + 336) + 2, column=5, value=value_B)

# Extract the dates from 33.xlsx
date_column_33 = ws_33['A']
date_row_2 = date_column_33[1].value  # Extract the date
date_row_29 = date_column_33[28].value  # Extract the date

# Save the workbook as 33.xlsx
filename = f"{date_row_2.strftime('%Y-%m-%d')}~{date_row_29.strftime('%Y-%m-%d')}.xlsx"
wb_33.save(filename)

# Check if the original 33.xlsx exists before renaming it
if os.path.exists('33.xlsx'):
    os.remove('33.xlsx')
    print("Original file 33.xlsx has been deleted.")

print(f"File {filename} has been created.")

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime, timedelta
import os

# Load workbook 11.xlsx and get Sheet3
wb_11 = load_workbook(filename='11.xlsx', data_only=True)
ws_11 = wb_11['Sheet3']

# Find the first Monday in the date column of Sheet3
date_column_11 = ws_11['A']
first_monday_row_11 = None
for cell in date_column_11:
    if isinstance(cell.value, datetime) and cell.value.weekday() == 0:  # Check if the cell contains a datetime object
        first_monday_row_11 = cell.row
        break

if first_monday_row_11 is None:
    raise ValueError("No Monday found in the date column of Sheet3 in 11.xlsx")

# Extract the dates from row 2 and row 29 of column A
date_row_2 = date_column_11[first_monday_row_11].value.date()  # Extract only the date part
date_row_29 = date_column_11[first_monday_row_11 + 27].value.date()  # Extract only the date part

# Create a new workbook 33.xlsx
wb_33 = Workbook()
ws_33 = wb_33.active
ws_33.title = 'Sheet1'

# Load workbook 22.xlsx and get Sheet1
wb_22 = load_workbook(filename='22.xlsx', read_only=True)
ws_22 = wb_22['Sheet1']

# Copy the whole Sheet1 of 22.xlsx to 33.xlsx with formatting
for row in ws_22.iter_rows(min_row=1, max_row=ws_22.max_row, min_col=1, max_col=ws_22.max_column):
    for cell in row:
        if cell.value is not None:
            copied_cell = ws_33.cell(row=cell.row, column=cell.column, value=cell.value)

            # Copy cell formatting
            if cell.has_style:
                copied_cell.font = Font(name=cell.font.name,
                                        size=cell.font.size,
                                        bold=cell.font.bold,
                                        italic=cell.font.italic,
                                        vertAlign=cell.font.vertAlign,
                                        underline=cell.font.underline,
                                        strike=cell.font.strike,
                                        color=cell.font.color)
                copied_cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                                  vertical=cell.alignment.vertical,
                                                  wrap_text=cell.alignment.wrap_text,
                                                  shrink_to_fit=cell.alignment.shrink_to_fit,
                                                  indent=cell.alignment.indent)
                copied_cell.number_format = cell.number_format
                copied_cell.protection = cell.protection
                copied_cell.border = cell.border
                copied_cell.fill = cell.fill

# Fill the first column of the new workbook from row 1 to row 28 with values from the second Monday
for i in range(first_monday_row_11 + 364, first_monday_row_11 + 392):
    date_value = ws_11.cell(row=i, column=1).value.date()  # Extract only the date part
    ws_33.cell(row=i - (first_monday_row_11 + 364) + 2, column=1, value=date_value)

    # Fill column E of the new workbook with values from column B of Sheet3 of 11.xlsx
    value_B = ws_11.cell(row=i, column=2).value
    ws_33.cell(row=i - (first_monday_row_11 + 364) + 2, column=5, value=value_B)

# Extract the dates from 33.xlsx
date_column_33 = ws_33['A']
date_row_2 = date_column_33[1].value  # Extract the date
date_row_29 = date_column_33[28].value  # Extract the date

# Save the workbook as 33.xlsx
filename = f"{date_row_2.strftime('%Y-%m-%d')}~{date_row_29.strftime('%Y-%m-%d')}.xlsx"
wb_33.save(filename)

# Check if the original 33.xlsx exists before renaming it
if os.path.exists('33.xlsx'):
    os.remove('33.xlsx')
    print("Original file 33.xlsx has been deleted.")

print(f"File {filename} has been created.")
